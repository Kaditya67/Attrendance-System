from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Max
from collections import defaultdict
import json
from django.urls import reverse  
from .models import Labs, Batches, Teacher, Student, Course, Attendance, Semester, LabsBatches
import openpyxl
from .forms import LabForm, TeacherUpdateForm
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER
from tkinter import CENTER

from myapp.studentViews import update_student

@login_required
def add_lab(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    
    if request.method == 'POST':
        form = LabForm(request.POST, teacher=teacher)
        if form.is_valid():
            form.save()
            return redirect('lab_dashboard')  # Redirect to some relevant page after saving
    else:
        form = LabForm(teacher=teacher)  # Pass teacher to the form

    return render(request, 'teachertemplates/add_lab.html', {'form': form, 'teacher': teacher})



@login_required
def lab_dashboard(request):

    if request.user.groups.filter(name='HOD').exists():
        is_hod = True
        is_principal = False
    elif request.user.groups.filter(name='Principal').exists():
        is_hod = False
        is_principal = True
    else:
        is_hod = False
        is_principal = False
        # Get the logged-in teacher
    teacher = get_object_or_404(Teacher, user=request.user)
    labs = teacher.assigned_labs.all()
    print(f"Assigned Labs: {labs}")
    
    if request.user.groups.filter(name="HOD").exists():
        labs = Labs.objects.all()
    print(f"Labs: {labs}")

    lab_data = []

    for lab in labs:
        lab_data.append({
            'lab_id': lab.id,
            'lab_name': lab.name,
            'lab_semester': lab.semester,
            'lab_index': lab.index,
        })
    print(f"Lab Data: {lab_data}")

    context = {
        'lab': lab_data,
        'labs': labs,  # Include all labs for navbar display
        'is_hod': is_hod,
        'is_principal': is_principal,
    }
    return render(request, 'teachertemplates/lab_dashboard.html', context)


def add_batches(request, lab_id):
    lab = get_object_or_404(Labs, id=lab_id)
    batch_instance, created = Batches.objects.get_or_create(lab=lab)
    
    if request.method == 'POST':
        batch_input = request.POST.get('batch_options')  # Example: "a,b,c"
        batch_list = [batch.strip() for batch in batch_input.split(',')]  # Split by commas
        batch_instance.set_batch_options(batch_list)
        return redirect('assign_batches', lab_id=lab.id)

    context = {
        'lab': lab,
        'batch_options': batch_instance.get_batch_options(),  # Retrieve the current batch options
    }
    return render(request, 'teachertemplates/add_batches.html', context)


def delete_batch(request, batch_id):
    if batch_id == 0:
        batch = Batches.objects.first()  # Or implement your own logic to find the right batch
    else:
        # Otherwise, retrieve the batch as normal
        batch = get_object_or_404(Batches, id=batch_id)
    
    lab_id = batch.lab.id  # Get the associated lab ID
    batch.delete()  # Delete the batch
    return redirect('lab_detail', lab_id=lab_id)  # Redirect to lab detail page


def assign_batches_to_students(request, lab_id):
    lab = get_object_or_404(Labs, id=lab_id)
    batch_obj = lab.batches.first()

    if batch_obj:
        batch_options = json.loads(batch_obj.batch_options)
    else:
        batch_options = []

    students = Student.objects.filter(semester=lab.semester)

    if request.method == 'POST':
        for student in students:
            selected_batch = request.POST.get(f'batch_{student.id}')
            if selected_batch:
                student.assign_batch(lab.index, selected_batch)

        # Redirect to the lab dashboard after assignment
        return redirect('lab_detail', lab_id=lab.id)

    return render(request, 'teachertemplates/assign_batches.html', {'lab': lab, 'batch_options': batch_options, 'students': students})


def lab_detail(request, lab_id):
    lab = get_object_or_404(Labs, id=lab_id)
    batches = lab.batches.all()  # Get all batches for the lab
    students = Student.objects.filter(semester=lab.semester) 
    print(students)

    index = lab.index

    for batch in batches:
        batch.batch_options = json.loads(batch.batch_options)  # Ensure it's a list

    batch_student_data = []
    # Iterate over each batch
    for batch in batches[0].batch_options:
        match_students = []
        for student in students:
            if(student.batches[str(index)] == batch):
                print(f'Batch: {batch}, Student: {student}')
                match_students.append({
                    'batch': batch,
                    'student': student
                })

        batch_student_data.append({
            'batch': batch,
            'match_students':match_students
        })
    
    print(f"Students and Batches {batch_student_data}")
        
    context = {
        'lab': lab,
        'batches': batches,  # Pass all batches
        'batch_student_data': batch_student_data
    }
    return render(request, 'teachertemplates/lab_detail.html', context)


from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages

def update_teacher(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)

    if request.user.groups.filter(name='HOD').exists():
        is_hod = True
        is_principal = False
    elif request.user.groups.filter(name='Principal').exists():
        is_hod = False
        is_principal = True
    else:
        is_hod = False
        is_principal = False

    
    if request.method == 'POST':
        if 'update_profile' in request.POST:
            form = TeacherUpdateForm(request.POST, instance=teacher)
            if form.is_valid():
                form.save()
                messages.success(request, 'Profile updated successfully!')
                return redirect('update_teacher', teacher_id=teacher.id)  # Redirect to the same page to show message

        elif 'change_password' in request.POST:
            # Handle password change logic here
            old_password = request.POST.get('old_password')
            new_password1 = request.POST.get('new_password1')
            new_password2 = request.POST.get('new_password2')

            # Add your password change logic here
            if new_password1 == new_password2 and teacher.user.check_password(old_password):
                teacher.user.set_password(new_password1)
                teacher.user.save()
                messages.success(request, 'Password changed successfully!')
                return redirect('update_teacher', teacher_id=teacher.id)  # Redirect to the same page to show message
            else:
                messages.error(request, 'Password change failed. Please check your inputs.')

    else:
        form = TeacherUpdateForm(instance=teacher)

    context = {
        'form': form,
        'teacher': teacher
    }
    return render(request, 'teachertemplates/update_teacher.html', context)



@login_required
def view_attendance(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    courses = teacher.assigned_courses.all()

    selected_course = None
    attendance_summary = []

    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        selected_course = get_object_or_404(Course, id=subject_id)

        # Fetch all attendance records for the selected course
        attendance_records = Attendance.objects.filter(course=selected_course)

        if not attendance_records.exists():
            messages.warning(request, "No attendance records found for the selected course.")
            return redirect('view_attendance')

        # Summarize attendance records by date and count
        summary = defaultdict(lambda: [0, '', 0])  # [total present, common notes, count]
        for record in attendance_records:
            # Only aggregate if the current count is greater than the previously saved count for that date
            summary[(record.date, record.count)][0] += record.present  # Total present
            summary[(record.date, record.count)][1] = record.notes  # Get notes
            summary[(record.date, record.count)][2] = record.count  # Use the last count found

        # Convert the summary to a list for rendering
        attendance_summary = [
            (date, total_present, notes, count) 
            for (date, count), (total_present, notes, _) in summary.items()
        ]

    context = {
        'courses': courses,
        'attendance_summary': attendance_summary,
        'selected_course': selected_course,
        'teacher': teacher
    }
    return render(request, 'teachertemplates/view_attendance.html', context)


@login_required
def update_Attendance(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    courses = teacher.assigned_courses.all()

    if request.method == 'POST':

        # Step 1: Fetch attendance records (Subject and Date selection)
        if 'fetch' in request.POST:
            subject_id = request.POST.get('subject')
            date = request.POST.get('date')

            course = get_object_or_404(Course, id=subject_id)
            attendance_records = Attendance.objects.filter(course=course, date=date)

            if not attendance_records.exists():
                messages.warning(request, "No attendance records found for the selected date.")
                return redirect('update_Attendance')

            max_count = attendance_records.aggregate(Max('count'))['count__max']
            lecture_numbers = list(range(1, max_count + 1))

            context = {
                'courses': courses,
                'lecture_numbers': lecture_numbers,
                'selected_course': course,
                'selected_date': date,
            }
            return render(request, 'teachertemplates/Update_Attedance.html', context)

        # Step 3: Update attendance records
        elif 'update_attendance' in request.POST and 'lecture_number' in request.POST:
            subject_id = request.POST.get('subject')
            date = request.POST.get('date')
            lecture_number = request.POST.get('lecture_number')

            course = get_object_or_404(Course, id=subject_id)
            attendance_records = Attendance.objects.filter(course=course, date=date, count=lecture_number)

            attendance_updated = 0
            for record in attendance_records:
                student_id = record.student.id
                attendance_status = request.POST.get(f'attendance_{student_id}', None)

                if attendance_status is not None:
                    record.present = (attendance_status == 'Present')
                    record.save()
                    attendance_updated += 1

            if attendance_updated > 0:
                messages.success(request, f"{attendance_updated} attendance records updated successfully!")
            else:
                messages.warning(request, "No attendance records were updated.")

            context = {
                'courses': courses,
                'attendance_records': attendance_records,
                'selected_course': course,
                'selected_date': date,
                'lecture_number': lecture_number,
            }
            return render(request, 'teachertemplates/Update_Attedance.html', context)
        # Step 2: Load attendance for a specific lecture number
        elif 'lecture_number' in request.POST:
            subject_id = request.POST.get('subject')
            date = request.POST.get('date')
            lecture_number = request.POST.get('lecture_number')

            course = get_object_or_404(Course, id=subject_id)
            attendance_records = Attendance.objects.filter(course=course, date=date, count=lecture_number)

            context = {
                'courses': courses,
                'attendance_records': attendance_records,
                'selected_course': course,
                'selected_date': date,
                'lecture_number': lecture_number,
            }
            return render(request, 'teachertemplates/Update_Attedance.html', context)
    else:
        context = {'courses': courses}
        return render(request, 'teachertemplates/Update_Attedance.html', context)


@login_required
def select_course_lecture(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    courses = teacher.assigned_courses.all()

    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        date = request.POST.get('date')

        course = get_object_or_404(Course, id=subject_id)
        attendance_records = Attendance.objects.filter(course=course, date=date)

        if not attendance_records.exists():
            messages.warning(request, "No attendance records found for the selected date.")
            return redirect('select_course_lecture')

        max_count = attendance_records.aggregate(Max('count'))['count__max']
        lecture_numbers = list(range(1, max_count + 1))

        context = {
            'courses': courses,
            'lecture_numbers': lecture_numbers,
            'selected_course': course,
            'selected_date': date,
            'teacher': teacher
        }

        # Redirect to edit attendance page after selecting the course and lecture
        lecture_number = request.POST.get('lecture_number')
        if lecture_number:
            return redirect(reverse('edit_attendance', args=[subject_id, date, lecture_number]))
        
        return render(request, 'teachertemplates/select_lecture.html', context)

    context = {'courses': courses, 'teacher': teacher}
    return render(request, 'teachertemplates/select_lecture.html', context)


@login_required
def edit_attendance(request, subject_id, date, lecture_number):
    teacher = get_object_or_404(Teacher, user=request.user)
    courses = teacher.assigned_courses.all()

    # Fetch the selected course
    course = get_object_or_404(Course, id=subject_id)

    # Fetch attendance records for the selected course, date, and lecture number
    attendance_records = Attendance.objects.filter(course=course, date=date, count=lecture_number)

    if not attendance_records.exists():
        messages.warning(request, "No attendance records found for the selected lecture.")
        return redirect('select_course_lecture')  # Redirect to course selection page if no records found

    if request.method == 'POST':
        attendance_updated = 0
        for record in attendance_records:
            student_id = record.student.id
            attendance_status = request.POST.get(f'attendance_{student_id}', None)

            if attendance_status is not None:
                record.present = (attendance_status == 'Present')
                record.save()
                attendance_updated += 1

        if attendance_updated > 0:
            messages.success(request, f"{attendance_updated} attendance records updated successfully!")
            return redirect('select_course_lecture')  # Redirect to selection page after update
        else:
            messages.warning(request, "No attendance records were updated.")

    context = {
        'courses': courses,
        'attendance_records': attendance_records,
        'selected_course': course,
        'selected_date': date,
        'lecture_number': lecture_number,
        'teacher': teacher
    }

    return render(request, 'teachertemplates/edit_attendance.html', context)


@login_required
def Add_Attendance(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    
    # Get all courses taught by the teacher
    courses = teacher.assigned_courses.all()

    course_data = []
    for course in courses:
        semester = course.semester
        students_in_semester = Student.objects.filter(semester=semester).distinct()

        course_data.append({
            'course': course,
            'semester': semester,
            'students': students_in_semester,
        })
    
    lab_batches = LabsBatches.objects.all()  # Get all lab batches
    
    context = {
        'course_data': course_data,
        'lab_batches': lab_batches,
        'teacher': teacher
    }

    return render(request, 'teachertemplates/add_attendance.html', context)


def get_teacher_courses(request):
    teacher = get_object_or_404(Teacher, user=request.user)

    # Get all courses assigned to the teacher
    courses = teacher.assigned_courses.all()

    course_data = []
    for course in courses:
        semester = course.semester  # Assuming course has a semester field
        students_in_semester = Student.objects.filter(semester=semester).distinct()  # Fetch distinct students for the semester

        # Append course, semester, and students information to course_data
        course_data.append({
            'course': course,
            'semester': semester,
            'students': students_in_semester,
        })

    return course_data


@login_required
def fetch_students(request):
    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        date = request.POST.get('date')

        # Get the selected course
        course = Course.objects.get(id=subject_id)
        semester = course.semester

        # Get students for the selected course's semester
        students = Student.objects.filter(semester=semester)

        context = {
            'students': students,
            'selected_course': course,
            'selected_date': date,
            'course_data': get_teacher_courses(request),  # Function to fetch teacher's courses
        }

        return render(request, 'teachertemplates/add_attendance.html', context)
    
    return redirect('Add_Attendance')


@login_required
def submit_attendance(request):
    if request.method == "POST":
        print("Received POST request for submit_attendance")
        subject_id = request.POST.get('subject')
        date = request.POST.get('date')
        common_notes = request.POST.get('common_notes', '')
        absent_students = request.POST.get('absent_students', '').split(',')
        
        print(f"Subject ID: {subject_id}, Date: {date}, Common Notes: {common_notes}, Absent Students: {absent_students}")
        course = Course.objects.get(id=subject_id)
        print(f"Course: {course}")
        # Retrieve current attendance records
        existing_attendance_records = Attendance.objects.filter(course=course, date=date)
        common_count = existing_attendance_records.last().count + 1 if existing_attendance_records.exists() else 1

        attendance_created = 0

        for key in request.POST:
            if key.startswith('attendance_'):
                student_id = key.split('_')[1]
                student = Student.objects.get(id=int(student_id))
                present = request.POST[key] == 'Present'

                Attendance.objects.create(
                    student=student,
                    course=course,
                    date=date,
                    present=present,
                    count=common_count,
                    notes=common_notes
                )
                attendance_created += 1

        for student_id in absent_students:
            if student_id:
                student = Student.objects.get(id=int(student_id))
                Attendance.objects.create(
                    student=student,
                    course=course,
                    date=date,
                    present=False,
                    count=common_count,
                    notes=common_notes
                )
                attendance_created += 1

        if attendance_created > 0:
            messages.success(request, f"{attendance_created} attendance records submitted successfully!")
        else:
            messages.warning(request, "No attendance records were created.")

        return redirect('Add_Attendance')

    return redirect('Add_Attendance')

@login_required
def Add_Lab_Attendance(request):
    teacher = get_object_or_404(Teacher, user=request.user)

    labs_data = []
    labs = teacher.assigned_labs.all()

    for lab in labs:
        lab_batches = Batches.objects.filter(lab=lab).values_list('batch_options', flat=True)
        
        # Deserialize `batch_options` if it contains JSON
        deserialized_batches = []
        for batch in lab_batches: 
            try:
                deserialized_batches.extend(json.loads(batch))   
            except json.JSONDecodeError:
                deserialized_batches.append(batch)   
         
        labs_data.append({
            'lab': lab.name,
            'batches': deserialized_batches,
        })
    
    context = {
        'teacher': teacher,
        'labs_data': labs_data
    }

    return render(request, 'teachertemplates/add_lab_attendance.html', context)

def get_labs_data_for_teacher(teacher):
    labs_data = []
    labs = teacher.assigned_labs.all()

    for lab in labs:
        lab_batches = Batches.objects.filter(lab=lab).values_list('batch_options', flat=True)
        
        # Deserialize `batch_options` if it contains JSON
        deserialized_batches = []
        for batch in lab_batches: 
            try:
                deserialized_batches.extend(json.loads(batch))   
            except json.JSONDecodeError:
                deserialized_batches.append(batch)   
         
        labs_data.append({
            'lab': lab.name,
            'batches': deserialized_batches,
        })
    return labs_data

@login_required
def fetch_lab_students(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    if request.method == 'POST':
        # print("Received POST request for fetch_lab_students :",request.POST)
        labName = request.POST.get('lab')
        lab_batch = request.POST.get('lab_batch')
        date = request.POST.get('date')

        # Get the selected lab batch
        lab = Labs.objects.get(name=labName)
        Semester = lab.semester
        temp_students = Student.objects.filter(semester=Semester)

        # Filter students according to batch
        Students = []
        for student in temp_students:
            if student.batches[str(lab.index)] == lab_batch:
                Students.append(student)
        students = Students
        
        # print("Students : ",students)
        context={
            'students': students,
            'selected_lab': lab.name,
            'selected_lab_batch': lab_batch,
            'selected_date': date,
            'labs_data': get_labs_data_for_teacher(teacher)
        }

        # print("Context : ",context)

        return render(request, 'teachertemplates/add_lab_attendance.html', context)

    return redirect('Add_Lab_Attendance')



from django.contrib import messages
from django.shortcuts import redirect
from .models import Labs, Student, LabAttendance
from datetime import datetime

@login_required
def submit_lab_attendance(request):
    if request.method == 'POST':
        lab_name = request.POST.get('lab')  # Lab name from the form
        lab_batch = request.POST.get('lab_batch')  # Batch name
        date_str = request.POST.get('date')  # Attendance date (string)
        common_notes = request.POST.get('common_notes', '')  # Optional notes
        absent_students = request.POST.get('absent_students', '').split(',')  # IDs of absent students

        # Print for debugging
        print(f"Request Post : ", request.POST)

        # Validate lab and batch
        try:
            lab = Labs.objects.get(name=lab_name)
        except Labs.DoesNotExist:
            messages.error(request, "Invalid lab selected.")
            return redirect('Add_Lab_Attendance')

        # Convert date from string to date object
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid date format.")
            return redirect('Add_Lab_Attendance')

        # Fetch students from the same semester as the lab
        try:
            semester = lab.semester
            temp_students = Student.objects.filter(semester=semester)
            students_in_batch = [
                student for student in temp_students
                if student.batches.get(str(lab.index)) == lab_batch
            ]
        except Exception as e:
            messages.error(request, f"Error fetching batch: {str(e)}")
            return redirect('Add_Lab_Attendance')

        # Ensure students exist for the selected batch
        if not students_in_batch:
            messages.warning(request, f"No students found for {lab_name}, batch {lab_batch}.")
            return redirect('Add_Lab_Attendance')

        # Prepare lab attendance records
        lab_attendance_records = []
        for student in students_in_batch:
            student_id = str(student.id)
            is_present = request.POST.get(f'attendance_{student_id}') == 'Present'
            is_absent = student_id in absent_students

            # Create lab attendance record
            lab_attendance_records.append(
                LabAttendance(
                    student=student,
                    lab=lab,
                    lab_batch=lab_batch,
                    date=date,
                    present=is_present and not is_absent,
                    notes=common_notes
                )
            )

        # Bulk create lab attendance records
        try:
            LabAttendance.objects.bulk_create(lab_attendance_records)
            messages.success(request, f"Lab attendance records for {len(lab_attendance_records)} students submitted successfully!")
        except Exception as e:
            messages.error(request, f"Error submitting lab attendance: {str(e)}")
            return redirect('Add_Lab_Attendance')

        # Redirect to the lab attendance page after successful submission
        return redirect('Add_Lab_Attendance')

    # Redirect to the attendance form if the request is not POST
    return redirect('Add_Lab_Attendance')

from collections import defaultdict
from django.contrib import messages
from django.shortcuts import render, get_object_or_404
from .models import Teacher, Labs, LabAttendance

@login_required
def view_lab_attendance(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    labs = teacher.assigned_labs.all()  # Assuming the teacher has assigned labs, similar to assigned_courses

    selected_lab = None
    lab_attendance_summary = []

    if request.method == 'POST':
        lab_id = request.POST.get('lab')
        selected_lab = get_object_or_404(Labs, id=lab_id)

        # Fetch all lab attendance records for the selected lab
        lab_attendance_records = LabAttendance.objects.filter(lab=selected_lab)

        if not lab_attendance_records.exists():
            messages.warning(request, "No lab attendance records found for the selected lab.")
            return redirect('view_lab_Attendance')

        # Summarize lab attendance records by date and count
        summary = defaultdict(lambda: [0, '', 0])  # [total present, common notes, count]
        for record in lab_attendance_records:
            # Only aggregate if the current count is greater than the previously saved count for that date
            summary[(record.date, record.lab_batch)][0] += record.present  # Total present
            summary[(record.date, record.lab_batch)][1] = record.notes  # Get notes
            summary[(record.date, record.lab_batch)][2] = record.lab_batch  # Use the last batch found

        # Convert the summary to a list for rendering
        lab_attendance_summary = [
            (date, total_present, notes, lab_batch) 
            for (date, lab_batch), (total_present, notes, _) in summary.items()
        ]

    context = {
        'labs': labs,
        'lab_attendance_summary': lab_attendance_summary,
        'selected_lab': selected_lab,
        'teacher': teacher
    }
    return render(request, 'teachertemplates/view_lab_attendance.html', context)

@login_required
def edit_lab_attendance(request, lab_id, date, batch):
    teacher = get_object_or_404(Teacher, user=request.user)
    labs = teacher.assigned_labs.all()  # Fetch the labs assigned to the teacher

    # Fetch the selected lab
    lab = get_object_or_404(Labs, id=lab_id)

    # Fetch attendance records for the selected lab, date, and batch
    lab_attendance_records = LabAttendance.objects.filter(lab=lab, date=date, lab_batch=batch)

    if not lab_attendance_records.exists():
        messages.warning(request, "No attendance records found for the selected lab and batch.")
        return redirect('select_lab')  # Redirect to lab selection page if no records found

    if request.method == 'POST':
        attendance_updated = 0
        for record in lab_attendance_records:
            student_id = record.student.id
            attendance_status = request.POST.get(f'attendance_{student_id}', None)

            if attendance_status is not None:
                record.present = (attendance_status == 'Present')
                record.save()
                attendance_updated += 1

        if attendance_updated > 0:
            messages.success(request, f"{attendance_updated} attendance records updated successfully!")
            return redirect('view_lab_Attendance')  # Redirect to lab selection page after update
        else:
            messages.warning(request, "No attendance records were updated.")

    context = {
        'labs': labs,
        'lab_attendance_records': lab_attendance_records,
        'selected_lab': lab,
        'selected_date': date,
        'batch': batch,
        'teacher': teacher
    }

    return render(request, 'teachertemplates/edit_lab_attendance.html', context)


from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch
# PDF Export Function with colors and formatting
def export_subject_attendance_pdf(course_data):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="subject_attendance_report.pdf"'

    # Create the PDF document
    pdf = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    # Define the styles for the document
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="TitleStyle",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        textColor=colors.darkblue,  # Title color
        fontSize=18
    )
    subtitle_style = ParagraphStyle(
        name="SubtitleStyle",
        parent=styles["Heading3"],
        alignment=TA_LEFT,
        textColor=colors.darkgreen,  # Subtitle color (for course names)
        fontSize=14
    )
    table_header_style = ParagraphStyle(
        name="TableHeaderStyle",
        parent=styles["Heading2"],
        textColor=colors.whitesmoke,
        alignment=TA_CENTER,
        fontSize=12,
        backColor=colors.grey
    )

    # Add title
    elements.append(Paragraph("Subject-wise Attendance Report", title_style))

    for course in course_data:
        # Add Course Title
        elements.append(Paragraph(f"Course: {course['course'].name}", subtitle_style))

        # Define table headers
        data = [["Student Name", "Total Classes", "Classes Attended", "Percentage"]]

        # Populate table with student attendance data for the course
        for student in course['students']:
            student_name = f"{student['student'].user.first_name} {student['student'].user.last_name}"
            data.append([
                student_name,
                len(student['attendance']),
                student['total_attended'],
                f"{student['percentage']}%"
            ])

        # Create and style the table
        table = Table(data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch, 1.2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align text
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
            ('FONTSIZE', (0, 0), (-1, 0), 12),  # Header font size
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Header padding
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),  # Row background color
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Grid color
            ('FONTSIZE', (0, 1), (-1, -1), 10),  # Row font size
            ('BACKGROUND', (0, 2), (-1, -1), colors.lightyellow),  # Alternating row colors
        ]))

        elements.append(table)
        elements.append(Paragraph("<br/><br/>", styles['Normal']))

    # Build and return the PDF
    pdf.build(elements)
    return response

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Excel Export Function with colors and formatting
def export_subject_attendance_excel(course_data):
    # Create a response object for Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="subject_attendance_report.xlsx"'

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subject Attendance Report"

    # Styles for headers and rows
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Grey fill
    row_fill1 = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue row
    row_fill2 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light yellow row

    alignment = Alignment(horizontal="center", vertical="center")

    # Add data for each course
    for course in course_data:
        # Add Course Title
        ws.append([f"Course: {course['course'].name}"])

        # Add headers
        headers = ["Student Name", "Total Classes", "Classes Attended", "Percentage"]
        ws.append(headers)

        # Apply header styles
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).font = header_font
            ws.cell(row=ws.max_row, column=col).fill = header_fill
            ws.cell(row=ws.max_row, column=col).alignment = alignment

        # Add student attendance data
        for idx, student in enumerate(course['students']):
            student_name = f"{student['student'].user.first_name} {student['student'].user.last_name}"
            total_classes = len(student['attendance'])
            total_attended = student['total_attended']
            percentage = student['percentage']
            row = [student_name, total_classes, total_attended, f"{percentage}%"]

            ws.append(row)

            # Apply alternating row colors
            fill = row_fill1 if idx % 2 == 0 else row_fill2
            for col in range(1, len(row) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill
                ws.cell(row=ws.max_row, column=col).alignment = alignment

        # Add a blank row after each course for clarity
        ws.append([])

    # Save the workbook to the response
    wb.save(response)
    return response


# Modify your original Subject_Attendance_Details view to handle export functionality
def Subject_Attendance_Details(request):
    # Check for HOD or Principal
    if request.user.groups.filter(name='Principal').exists():
        is_hod = False
        is_principal = True
    elif request.user.groups.filter(name='HOD').exists():
        is_hod = True
        is_principal = False
    else:
        is_hod = False
        is_principal = False

    # Get the teacher and courses
    teacher = get_object_or_404(Teacher, user=request.user)
    courses = teacher.assigned_courses.all()

    course_data = []
    attendance_data = {}
    defaulters_by_course = {}  # Dictionary to store defaulters by course

    for course in courses:
        semester = course.semester
        students_in_semester = Student.objects.filter(semester=semester).distinct()
        sample_student = students_in_semester.first()

        if sample_student:
            attendance_records = Attendance.objects.filter(course=course, student=sample_student).values('date', 'count').order_by('date')
            attendance_data[course.name] = []
            for record in attendance_records:
                attendance_data[course.name].append({'date': record['date'], 'count': record['count']})

            student_data = []
            student_percentages = []  # New array to store students and their percentages
            total_attended_sum = 0
            student_count = 0

            for student in students_in_semester:
                total_attended = 0
                total_classes = len(attendance_data[course.name])
                attendance_records = []
                for date_record in attendance_data[course.name]:
                    date = date_record['date']
                    count = date_record['count']
                    attendance_instance = Attendance.objects.filter(course=course, student=student, date=date, count=count).first()
                    is_present = attendance_instance.present if attendance_instance else None
                    attendance_records.append({
                        'date': date,
                        'latest_notes': attendance_instance.notes if attendance_instance else None,
                        'present': is_present,
                    })
                    if is_present:
                        total_attended += 1

                percentage = (total_attended / total_classes) * 100 if total_classes > 0 else 0
                student_data.append({
                    'student': student,
                    'attendance': attendance_records,
                    'total_attended': total_attended,
                    'percentage': round(percentage, 2),
                })
                
                # Add the student and their percentage to the new array
                student_percentages.append({
                    'student': student,
                    'percentage': round(percentage, 2),
                })

                total_attended_sum += total_attended
                student_count += 1

            average_attendance = (total_attended_sum / (student_count * total_classes)) * 100 if student_count > 0 and total_classes > 0 else 0
            # Determine the top 3 students based on percentage
            top_students = sorted(student_data, key=lambda x: x['percentage'], reverse=True)[:3]

            # Identify defaulters (students with attendance below 75%) for each course
            defaulters = [item for item in student_percentages if item['percentage'] < 75]

            course_data.append({
                'course': course,
                'semester': semester,
                'students': student_data,
                'average_attendance': round(average_attendance, 2),
                'student_percentages': student_percentages,  # Include the new array
                'top_students': top_students,  # Keep the top students
            })
            
            # Store defaulters for this specific course
            defaulters_by_course[course.name] = defaulters 
            
    # Export functionality
    export_format = request.GET.get('format')
    if export_format == 'pdf':
        return export_subject_attendance_pdf(course_data)
    elif export_format == 'excel':
        return export_subject_attendance_excel(course_data)

    return render(request, 'teachertemplates/teacher_Dashboard.html', {
        'teacher': teacher,
        'course_data': course_data,
        'attendance_data': attendance_data,
        'defaulters_by_course': defaulters_by_course,  # Add defaulters by course to context
        'is_hod': is_hod,
        'is_principal': is_principal,
    })



# Export to PDF
def export_students_to_pdf(semester_data, semester): 
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="subjectwise_attendance_report.pdf"'

    # Create the PDF document
    pdf = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    # Define the styles for the document
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = TA_CENTER

    # Add title
    elements.append(Paragraph("Subject-wise Attendance Report", title_style))

    # Add semester title
    if semester_data:
        semester_title = semester
        semester_heading = Paragraph(f"Semester: {semester_title}", styles['Heading2'])
        elements.append(semester_heading)

    # Loop through each subject (course) and compile a table for each
    for course in semester_data[0]['student_data'][0]['course_data']:
        course_name = course['course'].name
        elements.append(Paragraph(f"Subject: {course_name}", styles['Heading3']))

        # Define table headers
        data = [["Student Name", "Total Classes", "Classes Attended", "Percentage"]]

        # Populate table with student attendance data for the specific course
        for student_data in semester_data[0]['student_data']:
            student = student_data['student']
            student_name = f"{student.user.first_name} {student.user.last_name}"
            
            # Find attendance info for the specific course
            course_info = next((c for c in student_data['course_data'] if c['course'].name == course_name), None)
            if course_info:
                data.append([
                    student_name,
                    course_info['total_count'],
                    course_info['count_present'],
                    f"{course_info['percentage']}%"
                ])

        # Create the table for this subject
        table = Table(data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch, 1.2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  
            ('FONTSIZE', (0, 0), (-1, 0), 12),  
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  
            ('FONTSIZE', (0, 1), (-1, -1), 10),  
        ]))

        elements.append(table)
        elements.append(Paragraph("<br/><br/>", styles['Normal']))

    # Build and return the PDF
    pdf.build(elements)
    return response



def export_students_to_excel(semester_data):
    # Create a response object for Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_report_by_subject.xlsx"'

    # Create an Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Define a row for each course
    if semester_data:
        for semester in semester_data:
            for course in Course.objects.filter(semester=semester['semester']):
                # Add headers per course
                ws.append([f"Course: {course.name}"])
                headers = ["Student Name", "Total Classes", "Classes Attended", "Percentage"]
                ws.append(headers)

                # Populate data for each student in this course
                for student_data in semester['student_data']:
                    student = student_data['student']
                    student_name = f"{student.user.first_name} {student.user.last_name}"

                    # Get the attendance data for this specific course
                    course_data = next((cd for cd in student_data['course_data'] if cd['course'] == course), None)
                    if course_data:
                        total_count = course_data['total_count']
                        count_present = course_data['count_present']
                        percentage = course_data['percentage']
                        row = [student_name, total_count, count_present, f"{percentage}%"]
                        ws.append(row)

                # Add a blank row after each course for clarity
                ws.append([])

    # Save the workbook to the response
    wb.save(response)
    return response



# Class Report View with Export Functionality
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse

# Class Report View with Export Functionality
def Class_Report(request):
    teacher = get_object_or_404(Teacher, user=request.user)
    semesters = Semester.objects.filter(session_year=teacher.assigned_courses.first().semester.session_year)

    selected_semester = request.GET.get('semester')
    export_format = request.GET.get('format')  # Check if an export format (excel or pdf) is requested
    semester_data = []

    # If no semester is selected, return an error message or prompt
    if not selected_semester or selected_semester == 'None':
        # Show an error message on the page instead of throwing an error
        return render(request, 'teachertemplates/class_report.html', {
            'teacher': teacher,
            'semester_data': semester_data,
            'semesters': semesters,
            'selected_semester': None,
            'error_message': 'Please select a semester before exporting the report.'
        })

    # Try to convert selected_semester to integer (just in case)
    try:
        semester = get_object_or_404(Semester, id=int(selected_semester))
    except (ValueError, Semester.DoesNotExist):
        return HttpResponse("Invalid semester selected. Please select a valid semester.", status=400)

    # Now handle the case where a valid semester is selected
    students = Student.objects.filter(semester=semester)
    courses = Course.objects.filter(semester=semester)

    student_data_list = []
    for student in students:
        course_data_list = []
        total_present = 0
        total_classes = 0

        for course in courses:
            attendances = Attendance.objects.filter(student=student, course=course)
            total_count = attendances.count()
            count_present = attendances.filter(present=True).count()
            percentage = (count_present / total_count * 100) if total_count else 0

            course_data_list.append({
                'course': course,
                'total_count': total_count,
                'count_present': count_present,
                'percentage': round(percentage, 2),
            })

            total_present += count_present
            total_classes += total_count

        student_data_list.append({
            'student': student,
            'course_data': course_data_list,
            'total_present': total_present,
            'total_classes': total_classes,
        })

    semester_data.append({
        'semester': semester,
        'student_data': student_data_list,
    })

    # Handle export functionality
    if export_format == 'excel':
        return export_students_to_excel(semester_data)
    elif export_format == 'pdf':
        return export_students_to_pdf(semester_data,semester)

    return render(request, 'teachertemplates/class_report.html', {
        'teacher': teacher,
        'semester_data': semester_data,
        'semesters': semesters,
        'selected_semester': selected_semester
    })